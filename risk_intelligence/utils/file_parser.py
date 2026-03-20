"""
Parses uploaded project files (JSON, PDF, Excel, CSV) into structured data.

- JSON: Expected to match ProjectPlanInput schema directly.
- Excel/CSV: Expected columns — task_name, start_date, end_date, current_status, etc.
             Supports multiple sheets: 'tasks', 'resources', 'dependencies'.
- PDF: Extracts text, then uses the LLM to parse it into structured format.
"""

import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))

import json
import csv
import io
import logging

logger = logging.getLogger(__name__)


# ── JSON Parser ─────────────────────────────────────────────────────

def parse_json(file_bytes: bytes) -> dict:
    """Parse a JSON file directly into project plan dict."""
    raw = json.loads(file_bytes.decode("utf-8"))
    return raw


# ── CSV Parser ──────────────────────────────────────────────────────

def parse_csv(file_bytes: bytes) -> dict:
    """
    Parse a CSV file. Expects at minimum these columns:
    task_name, start_date, end_date, current_status

    Optional columns:
    description, resource_type, needed, currently_used, unit,
    dependency_name, dependency_type, dep_status
    """
    text = file_bytes.decode("utf-8")
    reader = csv.DictReader(io.StringIO(text))
    rows = [row for row in reader]

    if not rows:
        raise ValueError("CSV file is empty.")

    # Normalize column names (lowercase, strip spaces)
    rows = [{k.strip().lower().replace(" ", "_"): v.strip() for k, v in row.items()} for row in rows]

    tasks = []
    resources = []
    dependencies = []

    for row in rows:
        # Tasks
        if row.get("task_name"):
            tasks.append({
                "task_name": row["task_name"],
                "start_date": row.get("start_date", "2025-01-01"),
                "end_date": row.get("end_date", "2025-12-31"),
                "current_status": row.get("current_status", "not_started"),
                "description": row.get("description", ""),
            })

        # Resources (if resource columns exist)
        if row.get("resource_type"):
            resources.append({
                "resource_type": row["resource_type"],
                "needed": float(row.get("needed", 0)),
                "currently_used": float(row.get("currently_used", 0)),
                "unit": row.get("unit", "count"),
            })

        # Dependencies (if dependency columns exist)
        if row.get("dependency_name"):
            dependencies.append({
                "dependency_name": row["dependency_name"],
                "dependency_type": row.get("dependency_type", "internal"),
                "status": row.get("dep_status", row.get("dependency_status", "pending")),
                "blocking_tasks": [t.strip() for t in row.get("blocking_tasks", "").split(",") if t.strip()],
            })

    project_name = rows[0].get("project_name", "Uploaded Project")

    return {
        "project_name": project_name,
        "project_description": rows[0].get("project_description", ""),
        "tasks": tasks,
        "resources": resources,
        "dependencies": dependencies,
    }


# ── Excel Parser ────────────────────────────────────────────────────

def parse_excel(file_bytes: bytes) -> dict:
    """
    Parse an Excel file. Supports two modes:
    1. Multi-sheet: Separate sheets named 'tasks', 'resources', 'dependencies'
    2. Single-sheet: All data in one sheet (same columns as CSV)
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("Install openpyxl: pip install openpyxl")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    sheet_names_lower = {s.lower(): s for s in wb.sheetnames}

    project_name = "Uploaded Project"
    project_description = ""
    tasks = []
    resources = []
    dependencies = []

    def sheet_to_dicts(sheet):
        """Convert a sheet into a list of dicts using first row as headers."""
        data = []
        headers = None
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(h).strip().lower().replace(" ", "_") if h else f"col_{j}" for j, h in enumerate(row)]
                continue
            if all(v is None for v in row):
                continue
            data.append({headers[j]: (str(v).strip() if v is not None else "") for j, v in enumerate(row) if j < len(headers)})
        return data

    # Multi-sheet mode
    if "tasks" in sheet_names_lower:
        task_rows = sheet_to_dicts(wb[sheet_names_lower["tasks"]])
        for row in task_rows:
            if row.get("task_name"):
                tasks.append({
                    "task_name": row["task_name"],
                    "start_date": row.get("start_date", "2025-01-01"),
                    "end_date": row.get("end_date", "2025-12-31"),
                    "current_status": row.get("current_status", "not_started"),
                    "description": row.get("description", ""),
                })
                if not project_name or project_name == "Uploaded Project":
                    project_name = row.get("project_name", project_name)

        if "resources" in sheet_names_lower:
            res_rows = sheet_to_dicts(wb[sheet_names_lower["resources"]])
            for row in res_rows:
                if row.get("resource_type"):
                    resources.append({
                        "resource_type": row["resource_type"],
                        "needed": float(row.get("needed", 0)),
                        "currently_used": float(row.get("currently_used", 0)),
                        "unit": row.get("unit", "count"),
                    })

        if "dependencies" in sheet_names_lower:
            dep_rows = sheet_to_dicts(wb[sheet_names_lower["dependencies"]])
            for row in dep_rows:
                if row.get("dependency_name"):
                    dependencies.append({
                        "dependency_name": row["dependency_name"],
                        "dependency_type": row.get("dependency_type", "internal"),
                        "status": row.get("dep_status", row.get("dependency_status", row.get("status", "pending"))),
                        "blocking_tasks": [t.strip() for t in row.get("blocking_tasks", "").split(",") if t.strip()],
                    })

    else:
        # Single-sheet mode — treat first sheet like CSV
        first_sheet = wb[wb.sheetnames[0]]
        all_rows = sheet_to_dicts(first_sheet)

        for row in all_rows:
            if row.get("task_name"):
                tasks.append({
                    "task_name": row["task_name"],
                    "start_date": row.get("start_date", "2025-01-01"),
                    "end_date": row.get("end_date", "2025-12-31"),
                    "current_status": row.get("current_status", "not_started"),
                    "description": row.get("description", ""),
                })
            if row.get("resource_type"):
                resources.append({
                    "resource_type": row["resource_type"],
                    "needed": float(row.get("needed", 0)),
                    "currently_used": float(row.get("currently_used", 0)),
                    "unit": row.get("unit", "count"),
                })
            if row.get("dependency_name"):
                dependencies.append({
                    "dependency_name": row["dependency_name"],
                    "dependency_type": row.get("dependency_type", "internal"),
                    "status": row.get("dep_status", row.get("status", "pending")),
                    "blocking_tasks": [t.strip() for t in row.get("blocking_tasks", "").split(",") if t.strip()],
                })

        if all_rows:
            project_name = all_rows[0].get("project_name", project_name)
            project_description = all_rows[0].get("project_description", "")

    wb.close()

    return {
        "project_name": project_name,
        "project_description": project_description,
        "tasks": tasks,
        "resources": resources,
        "dependencies": dependencies,
    }


# ── PDF Parser (uses LLM to extract structured data) ────────────────

def parse_pdf(file_bytes: bytes, llm=None) -> dict:
    try:
        from pypdf import PdfReader
    except ImportError:
        raise ImportError("Install pypdf: pip install pypdf")

    reader = PdfReader(io.BytesIO(file_bytes))
    full_text = ""
    for page in reader.pages:
        full_text += page.extract_text() + "\n"

    full_text = full_text.strip()
    if not full_text:
        raise ValueError("Could not extract text from PDF.")

    if llm is None:
        return {
            "project_name": "Uploaded PDF Project",
            "project_description": full_text[:2000],
            "tasks": [{"task_name": "Review uploaded document", "start_date": "2025-01-01", "end_date": "2025-12-31", "current_status": "not_started"}],
            "resources": [],
            "dependencies": [],
        }

    prompt = f"""You are a Project Data Extractor. Your job is to extract REAL project management data from this document.

DOCUMENT TEXT:
{full_text[:6000]}

CRITICAL RULES:
1. Extract ONLY information that is EXPLICITLY stated or strongly implied in the document
2. DO NOT invent tasks, resources, or dependencies that aren't mentioned
3. If the document is NOT a project plan (e.g., it's a research paper, report, or analysis), then:
   - Set project_name to the document's title
   - Create tasks based on the actual phases/sections described
   - For resources, extract ONLY what is mentioned (team size, tools, budget)
   - For dependencies, extract ONLY what is mentioned (external tools, APIs, datasets, approvals)
4. If a field cannot be determined from the document, use EMPTY arrays — do NOT make up data
5. For current_status: 
   - If the document describes completed work → "completed"
   - If it describes ongoing work → "in_progress"  
   - If it describes future/planned work → "not_started"
   - If it mentions delays or issues → "delayed" or "blocked"
6. For resources:
   - ONLY include resources explicitly mentioned with quantities
   - If the document says "team of 5 engineers" → needed=5
   - If no quantity is mentioned, DO NOT include that resource
   - For budget, ONLY include if a specific dollar/rupee amount is mentioned
7. For dependencies:
   - ONLY include external systems, tools, datasets, or approvals mentioned
   - If something is described as unavailable or pending → status="blocked" or "pending"
   - If something is described as set up and working → status="resolved"

Respond ONLY with valid JSON:
{{
  "project_name": "Exact title from the document",
  "project_description": "1-2 sentence summary of what this document describes",
  "tasks": [
    {{
      "task_name": "Actual task/phase from the document",
      "start_date": "YYYY-MM-DD (use document dates if available, otherwise estimate)",
      "end_date": "YYYY-MM-DD",
      "current_status": "completed | in_progress | not_started | delayed | blocked",
      "description": "What this task involves, from the document"
    }}
  ],
  "resources": [
    {{
      "resource_type": "Only resources explicitly mentioned",
      "needed": 0,
      "currently_used": 0,
      "unit": "count | USD | hours"
    }}
  ],
  "dependencies": [
    {{
      "dependency_name": "Only dependencies explicitly mentioned",
      "dependency_type": "internal | external",
      "status": "pending | resolved | blocked",
      "blocking_tasks": []
    }}
  ]
}}

REMEMBER: Empty arrays are BETTER than invented data. If you're not sure, leave it out.
"""

    response = llm.invoke(prompt)
    raw = response.content.strip()

    if raw.startswith("```"):
        raw = raw.split("\n", 1)[-1]
    if raw.endswith("```"):
        raw = raw.rsplit("```", 1)[0]

    try:
        parsed = json.loads(raw.strip())
    except json.JSONDecodeError:
        logger.warning("LLM could not parse PDF into structured data, using raw text.")
        return {
            "project_name": "Uploaded PDF Project",
            "project_description": full_text[:2000],
            "tasks": [{"task_name": "Review uploaded document", "start_date": "2025-01-01", "end_date": "2025-12-31", "current_status": "not_started"}],
            "resources": [],
            "dependencies": [],
        }

    # POST-VALIDATION: Remove any resources with needed=0 or suspiciously generic data
    if "resources" in parsed:
        parsed["resources"] = [
            r for r in parsed["resources"]
            if r.get("needed", 0) > 0 and r.get("resource_type", "").strip() != ""
        ]

    # Ensure at least one task exists
    if not parsed.get("tasks"):
        parsed["tasks"] = [{
            "task_name": "Project Execution",
            "start_date": "2025-01-01",
            "end_date": "2025-12-31",
            "current_status": "in_progress",
            "description": parsed.get("project_description", "")
        }]

    return parsed


# ── Main Router ─────────────────────────────────────────────────────

def parse_upload(filename: str, file_bytes: bytes, llm=None) -> dict:
    """Route to the correct parser based on file extension."""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    if ext == "json":
        return parse_json(file_bytes)
    elif ext == "csv":
        return parse_csv(file_bytes)
    elif ext in ("xlsx", "xls"):
        return parse_excel(file_bytes)
    elif ext == "pdf":
        return parse_pdf(file_bytes, llm=llm)
    else:
        raise ValueError(f"Unsupported file format: .{ext}. Use JSON, CSV, Excel, or PDF.")
